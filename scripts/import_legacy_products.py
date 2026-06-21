#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REMOTE_HOST = "root@178.104.6.190"
REMOTE_SSH_KEY = Path("/Users/abcd/.ssh/id_ed25519")
REMOTE_SITE_ROOT = "/var/www/legacy-catalog"
REMOTE_IMAGE_DIR = f"{REMOTE_SITE_ROOT}/assets/product/Uploads/pro"
REMOTE_WORKDIR = "/root/global-castle-import"
MAX_IMAGE_EDGE = 1800
JPEG_QUALITY = "85"

CATEGORY_RULES = {
    "Stainless Steel Tumblers": {
        "top_category": "Steel",
        "sub_category": "Tumbler",
        "primary_path": ["Steel", "Tumbler"],
        "source_url": "https://www.cnurbaneco.com/list-5-78.html",
    },
    "Stainless Steel Bottles": {
        "top_category": "Steel",
        "sub_category": "Vacuum Bottle",
        "primary_path": ["Steel", "Vacuum Bottle"],
        "source_url": "https://www.cnurbaneco.com/list-5-35.html",
    },
    "Plastic Cup & Tumbler": {
        "top_category": "Plastic",
        "sub_category": "Tumbler",
        "primary_path": ["Plastic", "Tumbler"],
        "source_url": "https://www.cnurbaneco.com/list-5-59.html",
    },
}

REMOTE_APPLY_SCRIPT = r"""
import json
import re
from collections import Counter
from pathlib import Path

SITE_ROOT = Path("/var/www/legacy-catalog")
INDEX_PATH = SITE_ROOT / "index.html"
DATA_PATH = SITE_ROOT / "data.json"
LISTING_PATH = SITE_ROOT / "listing-data.json"
MANIFEST_PATH = Path("/root/global-castle-import/manifest.json")

manifest = json.loads(MANIFEST_PATH.read_text())
data = json.loads(DATA_PATH.read_text())
listing = json.loads(LISTING_PATH.read_text())
backup_tag = manifest["backup_tag"]

for path in [DATA_PATH, LISTING_PATH]:
    backup = path.with_suffix(path.suffix + f".bak-{backup_tag}")
    if not backup.exists():
        backup.write_text(path.read_text())

used_slugs = {item.get("slug") for item in data["products"] if item.get("slug")}
code_to_data = {item.get("product_item"): item for item in data["products"] if item.get("product_item")}
code_to_listing = {item.get("product_item"): item for item in listing["products"] if item.get("product_item")}


def unique_slug(base: str) -> str:
    slug = base
    counter = 2
    while slug in used_slugs:
        slug = f"{base}-{counter}"
        counter += 1
    used_slugs.add(slug)
    return slug


updated_codes = []
for rec in manifest["products"]:
    code = rec["item"]
    existing = code_to_data.get(code)
    existing_listing = code_to_listing.get(code)
    if existing:
        slug = existing.get("slug") or rec["slug"]
    else:
        slug = unique_slug(rec["slug"])

    full_item = {
        "slug": slug,
        "source_url": rec["source_url"],
        "title": rec["title"],
        "product_item": code,
        "category": rec["sub_category"],
        "top_category": rec["top_category"],
        "sub_category": rec["sub_category"],
        "branch_categories": [],
        "primary_path": rec["primary_path"],
        "category_paths": [rec["primary_path"]],
        "description_text": rec["description_text"],
        "main_image": rec["images"][0],
        "images": rec["images"],
        "source_status": (existing or {}).get("source_status") or "full",
        "uploaded_at": rec["uploaded_at"],
    }
    listing_item = {
        "slug": slug,
        "title": rec["title"],
        "product_item": code,
        "category": rec["sub_category"],
        "top_category": rec["top_category"],
        "sub_category": rec["sub_category"],
        "primary_path": rec["primary_path"],
        "category_paths": [rec["primary_path"]],
        "description_text": rec["description_text"],
        "main_image": rec["images"][0],
        "image_count": len(rec["images"]),
        "uploaded_at": rec["uploaded_at"],
    }

    if existing:
        existing.clear()
        existing.update(full_item)
    else:
        data["products"].append(full_item)

    if existing_listing:
        existing_listing.clear()
        existing_listing.update(listing_item)
    else:
        listing["products"].append(listing_item)

    updated_codes.append(code)

path_counter = Counter(tuple(item.get("primary_path", [])) for item in data["products"])


def walk(nodes):
    total = 0
    for node in nodes:
        path = tuple(node.get("path", []))
        if node.get("children"):
            child_total = walk(node["children"])
            node["count"] = child_total if child_total else path_counter.get(path, 0)
        else:
            node["count"] = path_counter.get(path, 0)
        total += node["count"]
    return total


walk(data.get("menu_tree", []))
walk(listing.get("menu_tree", []))

for collection in [data.get("categories", []), listing.get("categories", [])]:
    for cat in collection:
        path = tuple(cat.get("path", []))
        cat["count"] = path_counter.get(path, cat.get("count", 0))

data["generated_at"] = manifest["generated_at"]
listing["generated_at"] = manifest["generated_at"]
listing["total_products"] = len(listing["products"])

DATA_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n")
LISTING_PATH.write_text(json.dumps(listing, ensure_ascii=False, indent=2) + "\n")

text = INDEX_PATH.read_text()
sort_re = re.compile(
    r"function filteredItems\(\)\{const q=state\.query\.toLowerCase\(\); return state\.items\.filter"
    r"\(item=>\{if\(!productMatchesPath\(item,state\.path\)\) return false; if\(!q\) return true; "
    r"return \[item\.title,item\.product_item,item\.category,item\.top_category,item\.sub_category,"
    r"item\.description_text\]\.join\(' '\)\.toLowerCase\(\)\.includes\(q\);\}\);\}"
)
sort_replacement = (
    "function productSortValue(item){const time=Date.parse(item.uploaded_at||item.updated_at||''); "
    "return Number.isFinite(time)?time:0;}\n"
    "function filteredItems(){const q=state.query.toLowerCase(); return state.items.filter(item=>"
    "{if(!productMatchesPath(item,state.path)) return false; if(!q) return true; "
    "return [item.title,item.product_item,item.category,item.top_category,item.sub_category,"
    "item.description_text].join(' ').toLowerCase().includes(q);}).sort((a,b)=>productSortValue(b)-productSortValue(a));}"
)
if "function productSortValue(item)" not in text:
    text, count = sort_re.subn(sort_replacement, text, count=1)
    if count != 1:
        raise SystemExit("Could not patch index.html list sort logic")
    INDEX_PATH.write_text(text)

print(json.dumps({
    "updated_count": len(updated_codes),
    "updated_codes": updated_codes,
    "data_products": len(data["products"]),
    "listing_products": len(listing["products"]),
}))
"""


@dataclass
class ProductRecord:
    category: str
    item: str
    title: str
    description_text: str
    folder: Path
    top_category: str
    sub_category: str
    primary_path: list[str]
    source_url: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import batch products into the legacy catalog site.")
    parser.add_argument("--xlsx", required=True, type=Path, help="Excel file with the product list.")
    parser.add_argument("--products-root", required=True, type=Path, help="Root directory containing per-product folders.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and prepare files without uploading.")
    parser.add_argument(
        "--batch-tag",
        default=datetime.now().strftime("%Y%m%d"),
        help="Tag used in generated image filenames and backups. Default: current date in YYYYMMDD.",
    )
    parser.add_argument(
        "--ssh-key",
        type=Path,
        default=REMOTE_SSH_KEY,
        help=f"SSH private key for the legacy server. Default: {REMOTE_SSH_KEY}",
    )
    parser.add_argument(
        "--remote-host",
        default=REMOTE_HOST,
        help=f"SSH host for the legacy server. Default: {REMOTE_HOST}",
    )
    parser.add_argument(
        "--max-image-edge",
        type=int,
        default=MAX_IMAGE_EDGE,
        help=f"Maximum image edge in pixels after compression. Default: {MAX_IMAGE_EDGE}.",
    )
    return parser.parse_args()


def run(cmd: list[str], *, capture_output: bool = False) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        check=True,
        text=True,
        capture_output=capture_output,
    )


def load_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Workbook is empty: {xlsx_path}")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if rec.get("Catagory") and rec.get("Item") and rec.get("Name"):
            records.append(rec)
    return records


def normalize_number(value: Any) -> str | None:
    if value in (None, ""):
        return None
    try:
        as_float = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if abs(as_float - int(as_float)) < 1e-9:
        return str(int(as_float))
    return f"{as_float:.2f}".rstrip("0").rstrip(".")


def compact_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return "\n".join(part.strip() for part in str(value).splitlines() if part.strip())


def build_description(rec: dict[str, Any]) -> str:
    parts: list[str] = []
    features = compact_text(rec.get("Features"))
    material = compact_text(rec.get("Material"))
    lid = compact_text(rec.get("Lid"))
    cap_oz = normalize_number(rec.get("Capacity\noz"))
    cap_ml = normalize_number(rec.get("Capacity\nml"))
    size_w = normalize_number(rec.get("Size W"))
    size_l = normalize_number(rec.get("Size L"))
    size_h = normalize_number(rec.get("Size H"))

    if features:
        parts.append(f"Features:\n{features}")
    if material:
        parts.append(f"Material:\n{material}")
    if lid:
        parts.append(f"Lid:\n{lid}")
    capacity_bits = []
    if cap_oz:
        capacity_bits.append(f"{cap_oz} oz")
    if cap_ml:
        capacity_bits.append(f"{cap_ml} ml")
    if capacity_bits:
        parts.append(f"Capacity:\n{' / '.join(capacity_bits)}")
    if size_w or size_l or size_h:
        parts.append(f"Size:\nW {size_w or '-'} x L {size_l or '-'} x H {size_h or '-'} cm")
    return "\n\n".join(parts).strip()


def slugify(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", text.strip().lower())
    slug = re.sub(r"-+", "-", slug).strip("-")
    return slug or "product"


def locate_folder(products_root: Path, category: str, item: str) -> Path:
    direct = products_root / category / item
    if direct.is_dir():
        return direct

    exact_global = [path for path in products_root.glob(f"*/{item}") if path.is_dir()]
    if len(exact_global) == 1:
        return exact_global[0]
    if len(exact_global) > 1:
        raise ValueError(f"Multiple exact folder matches found for {item}: {exact_global}")

    prefix_matches = [path for path in products_root.glob(f"*/{item}*") if path.is_dir()]
    if len(prefix_matches) == 1:
        return prefix_matches[0]
    if len(prefix_matches) > 1:
        raise ValueError(f"Multiple prefix folder matches found for {item}: {prefix_matches}")

    raise FileNotFoundError(f"No folder found for {item} under {products_root}")


def image_sort_key(path: Path) -> tuple[int, str]:
    match = re.search(r"-(\d+)(?=\.[^.]+$)", path.name.replace(" ", ""))
    if match:
        return (int(match.group(1)), path.name.lower())
    return (9999, path.name.lower())


def build_products(records: list[dict[str, Any]], products_root: Path) -> list[ProductRecord]:
    built: list[ProductRecord] = []
    for rec in records:
        category = str(rec["Catagory"]).strip()
        item = str(rec["Item"]).strip()
        title = " ".join(str(rec["Name"]).split())
        if category not in CATEGORY_RULES:
            raise ValueError(f"Unsupported category in sheet: {category}")
        rule = CATEGORY_RULES[category]
        folder = locate_folder(products_root, category, item)
        built.append(
            ProductRecord(
                category=category,
                item=item,
                title=title,
                description_text=build_description(rec),
                folder=folder,
                top_category=rule["top_category"],
                sub_category=rule["sub_category"],
                primary_path=list(rule["primary_path"]),
                source_url=rule["source_url"],
            )
        )
    return built


def collect_images(folder: Path) -> list[Path]:
    images = [
        path
        for path in folder.iterdir()
        if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
    ]
    if not images:
        raise FileNotFoundError(f"No images found in {folder}")
    return sorted(images, key=image_sort_key)


def convert_image(src: Path, dest: Path, max_image_edge: int) -> None:
    cmd = [
        "/usr/bin/sips",
        "-s",
        "format",
        "jpeg",
        "-Z",
        str(max_image_edge),
        "--setProperty",
        "formatOptions",
        JPEG_QUALITY,
        str(src),
        "--out",
        str(dest),
    ]
    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def prepare_manifest(
    products: list[ProductRecord],
    batch_tag: str,
    max_image_edge: int,
    temp_root: Path,
) -> dict[str, Any]:
    image_dir = temp_root / "images"
    image_dir.mkdir(parents=True, exist_ok=True)

    generated_at = datetime.now(timezone(timedelta(hours=8))).replace(microsecond=0)
    manifest_products: list[dict[str, Any]] = []

    for offset, product in enumerate(products):
        uploaded_at = (generated_at + timedelta(minutes=offset)).isoformat()
        source_images = collect_images(product.folder)
        generated_images: list[str] = []
        for idx, source_image in enumerate(source_images, start=1):
            filename = f"{slugify(product.item)}-{batch_tag}-{idx}.jpg"
            convert_image(source_image, image_dir / filename, max_image_edge)
            generated_images.append(f"assets/product/Uploads/pro/{filename}")

        manifest_products.append(
            {
                "item": product.item,
                "title": product.title,
                "slug": slugify(f"{product.item}-{product.title}"),
                "category": product.category,
                "top_category": product.top_category,
                "sub_category": product.sub_category,
                "primary_path": product.primary_path,
                "source_url": product.source_url,
                "description_text": product.description_text,
                "images": generated_images,
                "uploaded_at": uploaded_at,
            }
        )

    manifest = {
        "generated_at": generated_at.isoformat(),
        "backup_tag": f"{batch_tag}-legacy-import",
        "products": manifest_products,
    }
    (temp_root / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n")
    return manifest


def ssh_base_args(args: argparse.Namespace) -> list[str]:
    return ["ssh", "-i", str(args.ssh_key), args.remote_host]


def scp_base_args(args: argparse.Namespace) -> list[str]:
    return ["scp", "-i", str(args.ssh_key)]


def remote_cmd(args: argparse.Namespace, command: str) -> list[str]:
    return ssh_base_args(args) + [command]


def upload_and_apply(args: argparse.Namespace, temp_root: Path) -> dict[str, Any]:
    manifest_path = temp_root / "manifest.json"
    script_path = temp_root / "apply_remote_import.py"
    script_path.write_text(REMOTE_APPLY_SCRIPT)

    run(remote_cmd(args, f"mkdir -p {REMOTE_WORKDIR} {REMOTE_IMAGE_DIR}"))

    images = sorted((temp_root / "images").glob("*.jpg"))
    if not images:
        raise RuntimeError("No generated images found to upload.")

    run(scp_base_args(args) + [str(path) for path in images] + [f"{args.remote_host}:{REMOTE_IMAGE_DIR}/"])
    run(scp_base_args(args) + [str(manifest_path), str(script_path), f"{args.remote_host}:{REMOTE_WORKDIR}/"])

    apply = run(
        remote_cmd(
            args,
            (
                f"python3 {REMOTE_WORKDIR}/apply_remote_import.py && "
                f"chmod 644 {REMOTE_IMAGE_DIR}/*-{args.batch_tag}-*.jpg"
            ),
        ),
        capture_output=True,
    )
    return json.loads(apply.stdout.strip().splitlines()[-1])


def print_summary(manifest: dict[str, Any], image_dir: Path) -> None:
    total_images = len(list(image_dir.glob("*.jpg")))
    print(f"Products prepared: {len(manifest['products'])}")
    print(f"Images prepared: {total_images}")
    for product in manifest["products"][:5]:
        print(f"- {product['item']}: {len(product['images'])} images, main={Path(product['images'][0]).name}")


def main() -> int:
    args = parse_args()
    if not args.xlsx.exists():
        raise FileNotFoundError(f"Excel file not found: {args.xlsx}")
    if not args.products_root.is_dir():
        raise NotADirectoryError(f"Products root not found: {args.products_root}")
    if not args.ssh_key.exists():
        raise FileNotFoundError(f"SSH key not found: {args.ssh_key}")

    rows = load_rows(args.xlsx)
    products = build_products(rows, args.products_root)

    with tempfile.TemporaryDirectory(prefix="legacy-import-") as temp_dir:
        temp_root = Path(temp_dir)
        manifest = prepare_manifest(products, args.batch_tag, args.max_image_edge, temp_root)
        print_summary(manifest, temp_root / "images")

        if args.dry_run:
            print("Dry run complete. No files were uploaded.")
            return 0

        result = upload_and_apply(args, temp_root)
        print("Remote import complete.")
        print(json.dumps(result, ensure_ascii=False, indent=2))

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
